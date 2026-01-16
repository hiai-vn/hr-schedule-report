"""
Node to export labeled messages from YAML to Excel file.

Input: labeled_messages dict from shared store
Output: Excel file with sheets for each category
"""
from pocketflow import Node
from datetime import datetime
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


class ExportExcelNode(Node):
    """Node to export labeled schedule messages to Excel.

    Input: labeled_messages dict with categories: nghi, tre, nua_buoi, remote
    Output: Excel file saved to output directory
    """

    # Category display names in Vietnamese
    CATEGORY_NAMES = {
        'nghi': 'Nghỉ phép',
        'tre': 'Đi trễ',
        'nua_buoi': 'Nghỉ nửa buổi',
        'remote': 'Làm remote',
    }

    # Header style
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    def prep(self, shared):
        """Get labeled messages from shared store."""
        return shared.get("labeled_messages", {})

    def exec(self, labeled_messages):
        """Create Excel workbook from labeled messages."""
        if not labeled_messages:
            return None

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Create a sheet for each category
        for category, display_name in self.CATEGORY_NAMES.items():
            messages = labeled_messages.get(category, [])
            self._create_category_sheet(wb, category, display_name, messages)

        # Create summary sheet
        self._create_summary_sheet(wb, labeled_messages)

        return wb

    def _create_category_sheet(self, wb, category, display_name, messages):
        """Create a sheet for a specific category."""
        ws = wb.create_sheet(title=display_name)

        # Headers
        headers = ['STT', 'Message ID', 'Tên', 'Ngày', 'Thông tin']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for idx, msg in enumerate(messages, 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=msg.get('message_id', ''))
            ws.cell(row=idx + 1, column=3, value=msg.get('name', ''))
            # Join dates with comma
            dates = msg.get('dates', [])
            dates_str = ', '.join(dates) if dates else ''
            ws.cell(row=idx + 1, column=4, value=dates_str)
            ws.cell(row=idx + 1, column=5, value=msg.get('info', ''))

        # Adjust column widths
        ws.column_dimensions['A'].width = 5   # STT
        ws.column_dimensions['B'].width = 12  # Message ID
        ws.column_dimensions['C'].width = 25  # Tên
        ws.column_dimensions['D'].width = 15  # Ngày
        ws.column_dimensions['E'].width = 40  # Thông tin

    def _create_summary_sheet(self, wb, labeled_messages):
        """Create summary sheet with counts."""
        ws = wb.create_sheet(title='Tổng hợp')
        # Move to first position
        wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)

        # Headers
        headers = ['Loại', 'Số lượng']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 2
        total = 0
        for category, display_name in self.CATEGORY_NAMES.items():
            count = len(labeled_messages.get(category, []))
            ws.cell(row=row, column=1, value=display_name)
            ws.cell(row=row, column=2, value=count)
            total += count
            row += 1

        # Total row
        ws.cell(row=row, column=1, value='Tổng cộng')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=2).font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12

    def post(self, shared, prep_res, exec_res):
        """Save workbook to file."""
        if not exec_res:
            return "default"

        # Create output directory if not exists
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"schedule_report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Save workbook
        exec_res.save(filepath)

        shared["excel_output_path"] = filepath
        print(f"Excel exported to: {filepath}")

        return "default"
